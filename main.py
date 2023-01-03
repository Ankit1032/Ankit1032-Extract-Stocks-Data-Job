
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import requests
import csv
import threading
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import datetime



def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
options = Options()
#options.add_experimental_option("detach",True)

# this parameter tells Chrome that
# it should be run without UI (Headless)
options.headless = True
#First time purpose
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)

browser = webdriver.Chrome(options=options)

browser.get("https://www.moneycontrol.com/stocksmarketsindia/")
browser.maximize_window() #Maximize the window

#obj = browser.find_element(By.CLASS_NAME, "robo_medium")
#print(obj)

# Beautiful Soup
# providing the page_source to soup | lxml -> parser
soup = BeautifulSoup(browser.page_source, 'lxml')
# finding relevant ids in page souce to extract shipment_id from the page
index_body = soup.find('table', class_='mctable1').find('tbody').find_all('tr')
#index_list = index_body.find('a',{'class':'robo_medium'}).text
print(len(index_body))


matrix = []
for i in range(0,len(index_body)):
    matrix_row = []
    matrix_row.append(index_body[i].find('a').text)
    matrix_row.append(index_body[i].find_all('td')[1].text)
    matrix_row.append(index_body[i].find_all('td')[2].text)
    matrix_row.append(index_body[i].find_all('td')[3].text)
    matrix.append(matrix_row)

#main code
# load workbook
workbook = Workbook()
worksheet = workbook.active

print(matrix)

# for i in range(len(matrix)):
#     for j in range(len(matrix[i])):
#         worksheet.cell(row=i+1, column=j+1).value = matrix[i][j]
#
# workbook.save("C:\Users\Ankit\Desktop\Jenkins-selenium automation\matrix.xlsx")

for i in range(len(matrix)):
    for j in range(len(matrix[i])):
        worksheet.cell(row=i+1, column=j+1).value = matrix[i][j]

current_date_time = datetime.datetime.now()
cur_sys_date = current_date_time.strftime('%Y-%m-%d')
cur_sys_time = current_date_time.strftime('%H-%M-%S')
cur_sys_dtime = cur_sys_date + "--" + cur_sys_time

path_file = "C:\\Users\\Ankit\\Desktop\\Jenkins-selenium automation\\stocks"+cur_sys_dtime+".xlsx"
print(path_file)
workbook.save(path_file)
print("Code Ends...")

#Add it to Jenkins